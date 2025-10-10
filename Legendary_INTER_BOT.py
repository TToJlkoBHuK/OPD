import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import time
import random

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "User-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache"
}

def request_with_retries(url, max_retries=3):
    for attempt in range(max_retries):
        try:
            response = requests.get(
                url,
                headers=headers,
                timeout=(10, 20),  # (connect timeout, read timeout)
                allow_redirects=True
            )
            if response.status_code == 200:
                return response
            print(f"Статус {response.status_code} для {url}, попытка {attempt+1}/{max_retries}")
        except requests.exceptions.RequestException as e:
            print(f"Ошибка запроса к {url}: {str(e)}, попытка {attempt+1}/{max_retries}")
        
        time.sleep(random.uniform(2, 5))  # Случайная задержка перед повтором
    return None

def check_last_online(admin_url):
    try:
        response = request_with_retries(admin_url)
        if not response:
            return 0
            
        soup = BeautifulSoup(response.text, 'html.parser')
        stats_div = soup.select_one('div.stats')
        if not stats_div:
            return 0
            
        time_tag = stats_div.find('time', datetime=True)
        if time_tag:
            last_online_str = time_tag['datetime'].rstrip('Z')
            last_online_time = datetime.fromisoformat(last_online_str)
            return 1 if (datetime.now() - last_online_time) <= timedelta(days=60) else 0
        return 0
    except Exception as e:
        print(f"Ошибка проверки активности {admin_url}: {str(e)}")
        return 0

def get_club_admins(club_url):
    response = request_with_retries(club_url)
    if not response:
        return []
    
    try:
        soup = BeautifulSoup(response.text, 'html.parser')
        admin_section = soup.select_one('section.team-show__meta')
        if not admin_section:
            return []
            
        return [
            (
                f'https://lichess.org{a["href"]}',
                a["href"].split('/@/')[-1]
            )
            for a in admin_section.select('a[href^="/@/"]')
        ]
    except Exception as e:
        print(f"Ошибка парсинга админов {club_url}: {str(e)}")
        return []

def collect_data():
    try:
        with open('clubs_domens.txt', 'r') as f:
            club_links = [line.strip() for line in f]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Клубы и админы"
        ws.append(["Клуб URL"] + [f"{col} {n}" for n in range(1, 11) for col in ["Админ URL", "Имя админа", "Статус"]])
        
        for idx, club_url in enumerate(club_links, 1):
            print(f"\n[{idx}/{len(club_links)}] Обработка: {club_url}")
            
            admins = get_club_admins(club_url)
            row = [club_url]
            
            for i, (admin_url, admin_name) in enumerate(admins[:10], 1):
                print(f"  Проверка админа {i}/{len(admins)}: {admin_name}")
                row.extend([admin_url, admin_name, check_last_online(admin_url)])
                time.sleep(random.uniform(1, 3))  # Случайная задержка между админами
            
            # Добавляем пустые значения для недостающих админов
            while len(row) < 31:
                row.extend(['', '', ''])
            
            ws.append(row[:31])
            wb.save("lichess_club_admins.xlsx")
            time.sleep(random.uniform(3, 7))  # Случайная задержка между клубами
        
        print("\nДанные успешно сохранены")
    
    except Exception as e:
        print(f"Критическая ошибка: {str(e)}")

def calculate_af_column():
    try:
        wb = load_workbook("lichess_club_admins.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=31):
            total = sum(cell.value or 0 for cell in row[3::3] if isinstance(cell.value, int))
            ws.cell(row=row[0].row, column=32, value=total)
        wb.save("lichess_club_admins.xlsx")
        print("Колонка AF обновлена")
    except Exception as e:
        print(f"Ошибка расчета AF: {str(e)}")

if __name__ == "__main__":
    collect_data()
    calculate_af_column()
