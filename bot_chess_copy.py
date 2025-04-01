import asyncio
import logging
from aiogram.utils.exceptions import MessageCantBeDeleted
from aiogram import Bot, Dispatcher, types
from aiogram.types import ParseMode, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from openpyxl import load_workbook
import os
import time
import shutil

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)
#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Инициализация бота
API_TOKEN = ''  # Замените на ваш токен
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)

# Файлы для сохранения данных
PROGRESS_FILE = "user_progress.txt"
BANS_FILE = "user_bans.txt"
USERS_FILE = "users.txt"  # Файл для хранения ID пользователей и их статусов
BROADCAST_TEMPLATE_FILE = "broadcast_template.txt"  # Файл для хранения шаблона рассылки
# Путь к папке для хранения медиа
MEDIA_FOLDER = "user_media"
GROUPS_DATA_FILE = "groups_data.txt"

# Глобальные переменные
ADMIN_IDS = [1881684121, 5312321185]  # 5312321185 Rus 1881684121
groups_data = []  # Все группы
sent_groups = {}  # Отправленные группы: {user_id: [group_index1, group_index2, ...]}
current_group_index = {}  # Текущая группа для каждого пользователя: {user_id: current_index}
user_bans = {}  # Блокировки пользователей: {user_id: ban_time}
users_status = {}  # Статусы пользователей: {user_id: status}
# Глобальная переменная для хранения никнеймов пользователей
user_nicknames = {}  # {user_id: nickname}
# Словарь для временного хранения медиагрупп
media_groups = {}
# Словарь для отслеживания времени последней отправки медиа
last_media_time = {}
# Очередь для отправки сообщений администраторам
message_queue = asyncio.Queue()

# Задержка между отправками сообщений (в секундах)
SEND_DELAY = 1

user_media_accumulation = {}
MEDIA_ACCUMULATION_DELAY = 2.0 # Секунды ожидания перед обработкой медиа

#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Создаем папку для медиа, если она не существует
if not os.path.exists(MEDIA_FOLDER):
    os.makedirs(MEDIA_FOLDER)

async def process_message_queue():
    while True:
        # Получаем сообщение из очереди
        chat_id, message_text, media = await message_queue.get()
        try:
            if media:
                # Если есть медиафайлы, отправляем медиагруппу
                await bot.send_media_group(chat_id=chat_id, media=media)
                logging.info(f"Медиагруппа отправлена админу {chat_id}.")
            else:
                # Если нет медиафайлов, отправляем текстовое сообщение
                await bot.send_message(chat_id=chat_id, text=message_text)
                logging.info(f"Текстовое сообщение отправлено админу {chat_id}.")
        except Exception as e:
            logging.error(f"Ошибка при отправке сообщения админу {chat_id}: {e}")
        finally:
            # Добавляем задержку между отправками
            await asyncio.sleep(SEND_DELAY)
            # Помечаем задачу как выполненную
            message_queue.task_done()

# Глобальная переменная для отслеживания текущей группы
global_group_index = 0

def load_data():
    global groups_data, sent_groups, current_group_index, user_bans, users_status, global_group_index, user_nicknames
    
    # Загрузка groups_data из файла
    groups_data = load_groups_data() or []
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) >= 2:  # Убедимся, что строка содержит хотя бы user_id и group_indices
                    user_id = parts[0]
                    group_indices = parts[1:]
                    try:
                        # Преобразуем user_id в int
                        user_id = int(user_id)
                        # Фильтруем только числовые значения для group_indices
                        valid_indices = [int(index) for index in group_indices if index.isdigit()]
                        sent_groups[user_id] = valid_indices
                        current_group_index[user_id] = len(valid_indices) - 1  # Последний индекс отправленной группы
                    except ValueError as e:
                        logging.error(f"Ошибка при загрузке данных из строки: {line.strip()}. Ошибка: {e}")
                        continue  # Пропускаем некорректные строки
        # Определяем максимальный индекс группы среди всех пользователей
        if sent_groups:  # Проверяем, что sent_groups не пуст
            max_group_index = max((max(indices) for indices in sent_groups.values()), default=0)
            global_group_index = max_group_index + 1  # Устанавливаем глобальный индекс на следующую группу
        else:
            global_group_index = 0  # Если sent_groups пуст, начинаем с нуля

    all_indices = set()
    duplicates_found = False
    for user_id, indices in sent_groups.items():
        for idx in indices:
            if idx in all_indices:
                logging.warning(f"Дубликат индекса {idx} у пользователя {user_id}")
                duplicates_found = True
            all_indices.add(idx)
    
    if duplicates_found:
        logging.error("Обнаружены дублирующиеся индексы в прогрессах пользователей!")
    
    if os.path.exists(BANS_FILE):
        with open(BANS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 2:  # Убедимся, что строка содержит user_id и ban_time
                    user_id, ban_time = parts
                    try:
                        user_bans[int(user_id)] = float(ban_time)
                    except ValueError as e:
                        logging.error(f"Ошибка при загрузке блокировки из строки: {line.strip()}. Ошибка: {e}")

    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",", 1)  # Разделяем по первой запятой
                if len(parts) == 2:  # Убедимся, что строка содержит user_id и status
                    user_id, status = parts
                    try:
                        users_status[int(user_id)] = status
                    except ValueError as e:
                        logging.error(f"Ошибка при загрузке статуса из строки: {line.strip()}. Ошибка: {e}")

    # Загрузка никнеймов
    if os.path.exists("nicknames.txt"):
        with open("nicknames.txt", "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(":")
                if len(parts) == 2:  # Убедимся, что строка содержит user_id и nickname
                    user_id, nickname = parts
                    try:
                        user_nicknames[int(user_id)] = nickname
                    except ValueError as e:
                        logging.error(f"Ошибка при загрузке никнейма из строки: {line.strip()}. Ошибка: {e}")
    if sent_groups:
        global_group_index = max(all_indices, default=0) + 1
    else:
        global_group_index = 0
# Сохранение данных в файлы
def save_data():
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        for user_id, group_indices in sent_groups.items():
            f.write(f"{user_id},{','.join(map(str, group_indices))}\n")
    with open(BANS_FILE, "w", encoding="utf-8") as f:
        for user_id, ban_time in user_bans.items():
            f.write(f"{user_id},{ban_time}\n")
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        for user_id, status in users_status.items():
            f.write(f"{user_id},{status}\n")
    # Сохранение никнеймов
    with open("nicknames.txt", "w", encoding="utf-8") as f:
        for user_id, nickname in user_nicknames.items():
            f.write(f"{user_id}:{nickname}\n")

# Проверка блокировки пользователя
def is_user_banned(user_id):
    if user_id in user_bans:
        # Проверяем, истекло ли время блокировки
        if time.time() < user_bans[user_id]:
            return True
        else:
            # Удаляем блокировку, если время истекло
            del user_bans[user_id]
            users_status[user_id] = ""  # Сбрасываем статус
            save_data()
    return False

def save_groups_data(groups):
    with open(GROUPS_DATA_FILE, "w", encoding="utf-8") as f:
        for group in groups:
            f.write(f"{group[0]},{group[1]}\n")

def load_groups_data():
    if os.path.exists(GROUPS_DATA_FILE):
        groups = []
        with open(GROUPS_DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 2:
                    groups.append((parts[0], int(parts[1])))
        return groups
    return None

#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Клавиатура для пользователей
def get_user_keyboard():
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("Получить следующую группу", callback_data="get_next_group"),
        InlineKeyboardButton("Получил бан", callback_data="ban_user")
    )
    # Добавляем новую кнопку "Шаблон рассылки"
    keyboard.row(
        InlineKeyboardButton("Шаблон рассылки", callback_data="show_broadcast_template")
    )
    # Добавляем кнопку "Инструкция"
    keyboard.row(
        InlineKeyboardButton("Инструкция", callback_data="send_instruction")
    )
    return keyboard

# Клавиатура для изменения статуса пользователя
def get_admin_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="❌", callback_data=f"set_status:{user_id}:❌"),
            InlineKeyboardButton(text="✅", callback_data=f"set_status:{user_id}:✅"),
            InlineKeyboardButton(text="Очистить", callback_data=f"set_status:{user_id}:")
        ],
        [InlineKeyboardButton(text="Снять бан", callback_data=f"remove_ban:{user_id}")],
        [InlineKeyboardButton(text="Просмотреть медиа", callback_data=f"view_media:{user_id}")],
        [InlineKeyboardButton(text="Вернуться к списку пользователей", callback_data="return_to_users_list")]
    ])

# Клавиатура для администраторов
def get_admin_panel():
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)  # Используем ReplyKeyboardMarkup
    keyboard.add(KeyboardButton("/run"), KeyboardButton("/send_groups"))
    keyboard.add(KeyboardButton("/users"), KeyboardButton("Изменить шаблон"))
    return keyboard

# Клавиатура для списка пользователей
def get_users_list_keyboard():
    keyboard = InlineKeyboardMarkup()
    for user_id, status in users_status.items():
        # Получаем никнейм пользователя или используем "ID: {user_id}", если никнейма нет
        nickname = user_nicknames.get(user_id, f"ID: {user_id}")
        if nickname.startswith("ID:"):
            button_text = f"{nickname} | Статус: {status or 'Без статуса'}"
        else:
            button_text = f"@{nickname} | Статус: {status or 'Без статуса'}"
        keyboard.add(InlineKeyboardButton(button_text, callback_data=f"user_select:{user_id}"))
    return keyboard

#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Загрузка шаблона рассылки
def load_broadcast_template():
    if os.path.exists(BROADCAST_TEMPLATE_FILE):
        with open(BROADCAST_TEMPLATE_FILE, "r", encoding="utf-8") as f:
            return f.read()
    return (
        "Hello! We invite your university team ([вставить название вуза в скобках]) to participate in the 8th Interuniversity Team Battle, "
        "the largest interuniversity online tournament on Lichess.\n\n"
        "📅 Date & Time: February 23, 2025 – 12:00 UTC\n"
        "⏳ Time Control: 3+0 Blitz, Rated\n"
        "🔗 Tournament Link: https://lichess.org/tournament/2OVexrSo\n\n"
        "1st Hunger Games – March 23, 2025 |Unique Format|\n"
        "https://lichess.org/tournament/oHZ8MI8c\n\n"
        "9th Interuniversity Team Battle – March 30, 2025 |3+3 Chess960|\n"
        "https://lichess.org/tournament/R9VQU47N\n\n"
        "10th Interuniversity Team Battle – April 27, 2025 |5+0 Blitz|\n"
        "https://lichess.org/tournament/dKuocHFV\n\n"
        "Participation is free and open to all university teams. Feel free to share this invitation with your club members. "
        "We look forward to seeing your team compete!\n\n"
        "Best regards,\n"
        "[вставить ваше имя]\n"
        "Interuniversity Team Battles Coordinator"
    )

# Сохранение шаблона рассылки
def save_broadcast_template(template_text):
    with open(BROADCAST_TEMPLATE_FILE, "w", encoding="utf-8") as f:
        f.write(template_text)

# Обработка Excel-файла и создание уникальных групп
def process_excel():
    logging.info("Processing Excel and forming groups...")
    filename = 'lichess_club_admins.xlsx'
    workbook = load_workbook(filename)
    sheet = workbook.active
    af_column_index = 32
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        club_url = row[0]
        active_admins = row[af_column_index - 1]
        data.append((club_url, active_admins))

    unique_data = list(set(data))
    unique_data_sorted = sorted(unique_data, key=lambda x: x[1], reverse=True)
    
    save_groups_data(unique_data_sorted)  # Сохраняем группы в файл
    return unique_data_sorted

# Функция для создания групп
def create_group(data_sorted, group_size, target_sum):
    current_group = []
    current_sum = 0
    removed_indices = []
    for i, (club_url, admins) in enumerate(data_sorted):
        if len(current_group) < group_size and current_sum + admins <= target_sum:
            current_group.append((club_url, admins))
            current_sum += admins
            removed_indices.append(i)
        if len(current_group) == group_size or current_sum >= target_sum:
            break
    data_sorted = [item for i, item in enumerate(data_sorted) if i not in removed_indices]
    return current_group, data_sorted

#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Автоматическое восстановление блокировки через 24 часа
async def restore_ban_if_inactive(user_id):
    await asyncio.sleep(24 * 60 * 60)  # Ждем 24 часа
    if user_id not in user_bans and user_id in users_status:
        # Проверяем, был ли пользователь активен за последние 24 часа
        if is_user_banned(user_id):
            return  # Пользователь уже заблокирован

        # Блокируем пользователя снова
        user_bans[user_id] = time.time() + 24 * 60 * 60  # Блокировка на 24 часа
        users_status[user_id] = ""  # Сбрасываем статус
        save_data()

        try:
            await bot.send_message(
                user_id,
                "Вы были автоматически заблокированы, так как не взаимодействовали с ботом в течение 24 часов."
            )
        except Exception as e:
            logging.error(f"Не удалось отправить сообщение пользователю {user_id}: {e}")

#===================================================================================================================================================================
#===================================================================================================================================================================
#===================================================================================================================================================================

# Команда /start
@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    user_id = message.from_user.id

    # Проверяем, является ли пользователь верифицированным
    if user_id not in ADMIN_IDS and user_id not in users_status:
        await message.reply("У вас нет доступа к этому боту.")
        return

    # Проверяем блокировку пользователя
    if is_user_banned(user_id):
        # Если пользователь заблокирован, отправляем сообщение и завершаем обработку
        await message.reply("Вы заблокированы. Вернитесь через 24 часа.")
        return

    # Логика для администраторов
    if user_id in ADMIN_IDS:
        await message.reply(
            "Привет, администратор! Вот ваша панель управления:",
            reply_markup=get_admin_panel()
        )
    else:
        # Для обычных пользователей
        keyboard = get_user_keyboard()
        await message.reply(
            "Привет! Нажмите кнопку ниже, чтобы получить группу.",
            reply_markup=keyboard
        )

# Команда /run (для администраторов)
@dp.message_handler(commands=['run'])
async def run_script(message: types.Message):
    global groups_data
    user_id = message.from_user.id
    if user_id not in ADMIN_IDS:
        await message.answer("У вас нет прав для выполнения этой команды.")
        return
    
    await message.answer("Запускаю скрипт... Пожалуйста, подождите.")
    try:
        # Если файл с группами уже существует, используем его
        if not os.path.exists(GROUPS_DATA_FILE):
            groups_data = process_excel()
        else:
            groups_data = load_groups_data()
        
        await message.answer(f"Группы успешно загружены. Всего групп: {len(groups_data)}")
    except Exception as e:
        logging.error(f"Error: {e}")
        await message.answer(f"Ошибка: {str(e)}")

# Команда /send_groups (для администраторов)
@dp.message_handler(commands=['send_groups'])
async def send_groups(message: types.Message):
    user_id = message.from_user.id
    if user_id not in ADMIN_IDS:
        await message.answer("У вас нет прав для выполнения этой команды.")
        return
    if not groups_data:
        await message.answer("Группы еще не созданы. Сначала выполните команду /run.")
        return
    await message.answer("Готово! Пользователи могут начать получать группы.")

#clean progress
@dp.message_handler(commands=['reset_groups'])
async def reset_groups(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    global sent_groups, groups_data
    sent_groups = {}
    groups_data = process_excel()  # Пересоздать группы
    save_data()
    await message.answer("Прогресс всех пользователей сброшен, группы пересозданы.")

# Команда /users (для администраторов)
@dp.message_handler(commands=['users'])
async def list_users(message: types.Message):
    user_id = message.from_user.id
    if user_id not in ADMIN_IDS:
        await message.answer("У вас нет прав для выполнения этой команды.")
        return
    if not users_status:
        await message.answer("Нет зарегистрированных пользователей.")
        return
    await message.answer("Список пользователей:", reply_markup=get_users_list_keyboard())
# Словарь для отслеживания состояний администраторов
admin_states = {}  # {user_id: state}

# Обработка кнопки "Изменить шаблон"
@dp.message_handler(lambda message: message.text == "Изменить шаблон")
async def change_broadcast_template(message: types.Message):
    user_id = message.from_user.id
    if user_id not in ADMIN_IDS:
        await message.answer("У вас нет прав для выполнения этого действия.")
        return
    # Переводим администратора в состояние ожидания нового шаблона
    admin_states[user_id] = "waiting_for_template"
    await message.answer("Пожалуйста, отправьте новый текст шаблона.")

# Обработка нового шаблона
@dp.message_handler(lambda message: message.from_user.id in admin_states and admin_states[message.from_user.id] == "waiting_for_template")
async def update_broadcast_template(message: types.Message):
    user_id = message.from_user.id
    new_template = message.text
    # Сохраняем новый шаблон в файл
    save_broadcast_template(new_template)
    # Удаляем состояние администратора
    del admin_states[user_id]
    await message.answer("Шаблон успешно обновлен!")

# Обработка кнопки "Шаблон рассылки"
@dp.callback_query_handler(lambda c: c.data == "show_broadcast_template")
async def show_broadcast_template(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    # Загружаем текущий шаблон
    template_text = load_broadcast_template()
    # Клавиатура с кнопкой возврата
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("Вернуться в меню", callback_data="return_to_menu")
    )
    # Отправляем шаблон текста и клавиатуру
    await bot.send_message(user_id, template_text, reply_markup=keyboard)
    # Удаляем старое сообщение с кнопками
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

# Обработка кнопки "Вернуться в меню"
@dp.callback_query_handler(lambda c: c.data == "return_to_menu")
async def return_to_menu(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    # Отправляем главное меню
    keyboard = get_user_keyboard()
    await bot.send_message(user_id, "Вы вернулись в главное меню.", reply_markup=keyboard)

    # Удаляем старое сообщение с кнопками
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

# Обработка кнопки "Инструкция"
@dp.callback_query_handler(lambda c: c.data == "send_instruction")
async def send_instruction(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    # Путь к файлу инструкции (предполагается, что файл лежит в папке проекта)
    instruction_file = "Инструкция.pdf"

    # Проверяем, существует ли файл
    if not os.path.exists(instruction_file):
        await bot.send_message(user_id, "Извините, инструкция временно недоступна.")
        return

    # Клавиатура с кнопкой возврата
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("Вернуться в меню", callback_data="return_to_menu")
    )

    # Отправляем PDF-файл
    with open(instruction_file, "rb") as file:
        await bot.send_document(
            chat_id=user_id,
            document=file,
            caption="Инструкция по использованию бота и рассылке.",
            reply_markup=keyboard
        )

    # Удаляем старое сообщение с кнопками
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

@dp.callback_query_handler(lambda c: c.data == "return_to_users_list")
async def return_to_users_list(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id

    if admin_id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет прав для выполнения этого действия.")
        return

    # Удаляем старое сообщение
    await bot.delete_message(chat_id=admin_id, message_id=callback_query.message.message_id)

    # Отправляем список пользователей
    await bot.send_message(admin_id, "Список пользователей:", reply_markup=get_users_list_keyboard())

# Обработка кнопки "Снять бан"
@dp.callback_query_handler(lambda c: c.data.startswith("remove_ban:"))
async def remove_ban(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    if admin_id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет прав для выполнения этого действия.")
        return
    
    _, user_id = callback_query.data.split(":")
    user_id = int(user_id)
    
    # Проверяем, заблокирован ли пользователь
    if user_id not in user_bans:
        await bot.answer_callback_query(callback_query.id, "Пользователь не заблокирован.")
        return
    
    # Снимаем блокировку
    del user_bans[user_id]
    save_data()
    
    # Отправляем уведомление пользователю
    try:
        await bot.send_message(
            user_id,
            "Администратор снял вашу блокировку. Вы можете продолжить получать группы.",
            reply_markup=get_user_keyboard()
        )
    except Exception as e:
        logging.error(f"Не удалось отправить сообщение пользователю {user_id}: {e}")
        await bot.answer_callback_query(callback_query.id, "Не удалось отправить уведомление пользователю.")
        return
    
    # Удаляем старое сообщение
    await bot.delete_message(chat_id=admin_id, message_id=callback_query.message.message_id)
    
    # Отправляем новое сообщение с подтверждением и кнопкой возврата
    await bot.send_message(
        chat_id=admin_id,
        text=f"Блокировка пользователя {user_id} успешно снята.",
        reply_markup=get_admin_keyboard(user_id)
    )
    
    # Запускаем таймер на 24 часа
    asyncio.create_task(restore_ban_if_inactive(user_id))

# Обработка кнопки "Просмотреть медиа"
@dp.callback_query_handler(lambda c: c.data.startswith("view_media:"))
async def view_media(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    _, user_id = callback_query.data.split(":")
    user_id = int(user_id)

    # Путь к папке пользователя
    user_folder = os.path.join(MEDIA_FOLDER, str(user_id))

    if os.path.exists(user_folder):
        # Формируем список медиафайлов
        media_files = []
        for file_name in os.listdir(user_folder):
            file_path = os.path.join(user_folder, file_name)
            if file_name.endswith(".jpg"):
                media_files.append(types.InputMediaPhoto(open(file_path, "rb")))
            elif file_name.endswith(".doc"):
                media_files.append(types.InputMediaDocument(open(file_path, "rb")))
            elif file_name.endswith(".mp4"):
                media_files.append(types.InputMediaVideo(open(file_path, "rb")))

        # Разделяем медиафайлы на группы по 10 элементов
        media_groups = [media_files[i:i + 10] for i in range(0, len(media_files), 10)]

        if media_groups:
            keyboard = InlineKeyboardMarkup()
            keyboard.row(
                InlineKeyboardButton("Вернуться", callback_data=f"user_select:{user_id}")
            )

            # Отправляем каждую группу медиа
            for group in media_groups:
                await bot.send_media_group(chat_id=admin_id, media=group)

            # Отправляем сообщение после всех медиа
            await bot.send_message(chat_id=admin_id, text="Медиа пользователя:", reply_markup=keyboard)
        else:
            await bot.send_message(chat_id=admin_id, text="У пользователя нет медиа.", reply_markup=get_admin_keyboard(user_id))
    else:
        # Если папка не существует, отправляем сообщение об этом
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton("Вернуться", callback_data=f"user_select:{user_id}")
        )
        await bot.send_message(chat_id=admin_id, text="У пользователя нет медиа.", reply_markup=keyboard)

# Обработка выбора пользователя
@dp.callback_query_handler(lambda c: c.data.startswith("user_select:"))
async def select_user(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    _, user_id = callback_query.data.split(":")
    user_id = int(user_id)
    await bot.edit_message_text(
        chat_id=admin_id,
        message_id=callback_query.message.message_id,
        text=f"Выбран пользователь ID: {user_id}. Измените его статус:",
        reply_markup=get_admin_keyboard(user_id)
    )

# Изменение статуса пользователя
@dp.callback_query_handler(lambda c: c.data.startswith("set_status:"))
async def set_user_status(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    if admin_id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет прав для выполнения этого действия.")
        return
    
    _, user_id, status = callback_query.data.split(":")
    user_id = int(user_id)
    users_status[user_id] = status
    save_data()
    
    # Удаляем старое сообщение
    await bot.delete_message(chat_id=admin_id, message_id=callback_query.message.message_id)
    
    # Отправляем новое сообщение с обновлённым статусом и кнопкой возврата
    await bot.send_message(
        chat_id=admin_id,
        text=f"Статус пользователя {user_id} изменен на: {status or 'Без статуса'}",
        reply_markup=get_admin_keyboard(user_id)
    )

# Обработка кнопки "Получить следующую группу"
@dp.callback_query_handler(lambda c: c.data == "get_next_group")
async def get_next_group(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    if not groups_data:
        await bot.answer_callback_query(callback_query.id, "Нет доступных групп.")
        return

    # Получаем список всех использованных индексов
    used_indices = set()
    for indices in sent_groups.values():
        used_indices.update(indices)

    # Ищем первый свободный индекс
    next_index = 0
    while next_index < len(groups_data):
        if next_index not in used_indices:
            break
        next_index += 1
    else:
        await bot.answer_callback_query(callback_query.id, "Все группы закончились.")
        return

    # Сохраняем прогресс
    if user_id not in sent_groups:
        sent_groups[user_id] = []
    sent_groups[user_id].append(next_index)
    save_data()

    # Отправляем группу
    club_url, active_admins = groups_data[next_index]
    await bot.send_message(user_id, f"Группа {next_index+1}/{len(groups_data)}:\n{club_url} ({active_admins} админов)")

    # Удаляем старое сообщение с кнопкой
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

    # Отправляем новую клавиатуру
    await bot.send_message(user_id, "Что дальше?", reply_markup=get_user_keyboard())

# Обработка кнопки "Получил бан"
@dp.callback_query_handler(lambda c: c.data == "ban_user")
async def ban_user(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    # Проверяем, зарегистрирован ли пользователь
    if user_id not in users_status:
        await bot.answer_callback_query(callback_query.id, "У вас нет доступа к этому действию.")
        return

    try:
        # Удаляем старое сообщение с кнопкой
        await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)
    except MessageCantBeDeleted:
        # Логируем ошибку, но продолжаем выполнение
        logging.warning(f"Не удалось удалить сообщение для пользователя {user_id}. Сообщение старше 48 часов или имеет другие ограничения.")

    # Блокируем пользователя на 24 часа
    ban_time = time.time() + 24 * 60 * 60  # 24 часа
    user_bans[user_id] = ban_time
    users_status[user_id] = ""  # Сбрасываем статус
    save_data()

    await bot.send_message(user_id, "Вы заблокированы. Вернитесь через 24 часа.")

# Словарь для временного хранения медиагрупп
media_groups = {}
# Словарь для отслеживания обработанных медиагрупп
processed_media_groups = {}
# Блокировка для предотвращения состояния гонки
media_groups_lock = asyncio.Lock()

# Обработка медиагрупп
async def process_accumulated_media(user_id: int):
    """
    Обрабатывает накопленные медиафайлы для пользователя:
    удаляет старые, сохраняет новые.
    """
    if user_id not in user_media_accumulation:
        return # Ничего не накопилось или уже обработано

    # Получаем данные и сразу удаляем из словаря, чтобы избежать повторной обработки
    media_data = user_media_accumulation.pop(user_id, None)
    if not media_data or not media_data.get("files"):
        return # Пустые данные

    messages_to_process = media_data["files"]
    user_folder = os.path.join(MEDIA_FOLDER, str(user_id))
    username = messages_to_process[0].from_user.username or f"ID: {user_id}" # Берем из первого сообщения

    logging.info(f"Начало обработки {len(messages_to_process)} медиафайлов для пользователя {username} ({user_id}).")

    # 1. Удаляем старые файлы (если папка существует)
    if os.path.exists(user_folder):
        try:
            shutil.rmtree(user_folder)
            logging.info(f"Старая папка {user_folder} удалена.")
        except Exception as e:
            logging.error(f"Ошибка при удалении старой папки {user_folder}: {e}")
            # Продолжаем выполнение, попытаемся создать папку заново

    # 2. Создаем папку заново
    try:
        os.makedirs(user_folder, exist_ok=True)
    except Exception as e:
        logging.error(f"Не удалось создать папку {user_folder}: {e}")
        # Отправляем сообщение об ошибке пользователю? Или только админу?
        # await bot.send_message(user_id, "Ошибка сохранения ваших файлов.")
        return # Прерываем обработку, если не можем создать папку

    # 3. Скачиваем и сохраняем новые файлы
    saved_count = 0
    for message in messages_to_process:
        file_info = None
        file_ext = ".dat" # Расширение по умолчанию
        file_name_base = str(time.time_ns()) # Уникальное имя файла на основе времени

        if message.photo:
            file_info = message.photo[-1] # Берем наибольшее разрешение
            file_ext = ".jpg"
            file_name_base = file_info.file_unique_id # Используем уникальный ID файла
        elif message.video:
            file_info = message.video
            file_ext = ".mp4"
            file_name_base = file_info.file_unique_id
        elif message.document:
            file_info = message.document
            # Пытаемся получить расширение из имени файла, если есть
            original_filename = getattr(file_info, 'file_name', '')
            if original_filename:
                 _, ext = os.path.splitext(original_filename)
                 if ext: file_ext = ext.lower()
            file_name_base = file_info.file_unique_id
        # Добавить другие типы медиа при необходимости (audio, voice, animation)

        if file_info:
            destination_path = os.path.join(user_folder, f"{file_name_base}{file_ext}")
            try:
                await bot.download_file_by_id(file_info.file_id, destination=destination_path)
                saved_count += 1
                # logging.debug(f"Файл {file_info.file_id} сохранен в {destination_path}")
            except Exception as e:
                logging.error(f"Ошибка скачивания файла {file_info.file_id} для пользователя {user_id}: {e}")
        else:
             logging.warning(f"Сообщение {message.message_id} от {user_id} не содержит известного типа медиа для сохранения.")


    logging.info(f"Завершение обработки медиа для {username} ({user_id}). Сохранено {saved_count} из {len(messages_to_process)} файлов.")
    # Можно отправить подтверждение пользователю (опционально)
    # try:
    #     await bot.send_message(user_id, f"✅ Ваши {saved_count} медиафайла(ов) сохранены.")
    # except Exception:
    #      pass

# --- Обновленный обработчик ЛЮБЫХ сообщений для накопления медиа ---
@dp.message_handler(content_types=types.ContentType.ANY)
async def handle_all_messages(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username

    # Сначала обновляем никнейм (если нужно)
    if username and user_nicknames.get(user_id) != username:
         user_nicknames[user_id] = username
         save_data()
         logging.info(f"Обновлен никнейм (через сообщение) для {user_id}: @{username}")
    elif user_id not in user_nicknames and not username: # Если ника нет и не было
         if user_id not in user_nicknames or user_nicknames[user_id] is None: # Проверяем, чтобы не перезаписать существующий пустой
             user_nicknames[user_id] = ""
             save_data()

    # Проверяем, содержит ли сообщение медиа, которое мы хотим сохранить
    is_media_to_save = message.photo or message.video or message.document # Добавьте другие типы если нужно

    if is_media_to_save:
        logging.debug(f"Получено медиа сообщение {message.message_id} от {user_id}")
        # Гарантируем наличие записи для пользователя
        if user_id not in user_media_accumulation:
            user_media_accumulation[user_id] = {"files": [], "timer_task": None}

        # Отменяем предыдущий таймер, если он есть
        if user_media_accumulation[user_id]["timer_task"]:
            user_media_accumulation[user_id]["timer_task"].cancel()
            logging.debug(f"Таймер обработки медиа для {user_id} отменен.")

        # Добавляем текущее сообщение в список
        user_media_accumulation[user_id]["files"].append(message)
        logging.debug(f"Медиа сообщение {message.message_id} добавлено в очередь для {user_id}. Всего в очереди: {len(user_media_accumulation[user_id]['files'])}")


        # Запускаем новый таймер
        # Создаем задачу, которая сначала ждет, а потом вызывает обработчик
        new_task = asyncio.create_task(
             delayed_media_processing(user_id, MEDIA_ACCUMULATION_DELAY)
             )
        user_media_accumulation[user_id]["timer_task"] = new_task
        logging.debug(f"Новый таймер обработки медиа для {user_id} запущен на {MEDIA_ACCUMULATION_DELAY} сек.")

    else:
        # Если это не медиа, возможно, это команда или текст, который должен обрабатываться другими хендлерами
        # Если этот хендлер последний, он поймает все, что не поймали предыдущие.
        # Можно добавить логирование или сообщение пользователю "Неизвестная команда"
        if not message.text or not message.text.startswith('/'): # Игнорируем команды, они должны быть обработаны выше
             logging.info(f"Получено не-медиа сообщение от {user_id}: {message.text[:50]}...")
             # await message.reply("Неизвестная команда или тип сообщения.") # Отвечать или нет - по желанию

async def delayed_media_processing(user_id: int, delay: float):
     """Корутина, которая ждет `delay` секунд, а затем вызывает обработчик."""
     await asyncio.sleep(delay)
     logging.debug(f"Таймер для {user_id} истек. Запуск process_accumulated_media.")
     try:
         # Запускаем фактическую обработку в фоне, чтобы не блокировать
         asyncio.create_task(process_accumulated_media(user_id))
     except Exception as e:
         logging.error(f"Ошибка при запуске process_accumulated_media для {user_id}: {e}")

# Обновление никнеймов при взаимодействии с ботом
@dp.message_handler()
async def update_user_nickname(message: types.Message):
    user_id = message.from_user.id
    nickname = message.from_user.username  # Берем username, если он существует
    if nickname:  # Добавляем ник только если он существует
        user_nicknames[user_id] = nickname

async def on_startup(dp):
    # Запускаем обработку очереди сообщений как фоновую задачу
    asyncio.create_task(process_message_queue())

# Запуск бота
if __name__ == '__main__':
    logging.info("Бот запущен")
    load_data()
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
