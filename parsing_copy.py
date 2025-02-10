import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime, timedelta

# Заголовки для избегания блокировок
headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "User-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
}

# Функция для проверки времени последнего визита
def check_last_online(admin_url):
    try:
        response = requests.get(admin_url, headers=headers)
        if response.status_code != 200:
            print(f"Не удалось получить данные для {admin_url}")
            return 0
        soup = BeautifulSoup(response.text, 'html.parser')
        stats_div = soup.find('div', class_='stats')
        if not stats_div:
            print(f"Не удалось найти блок с информацией для {admin_url}")
            return 0
        time_tag = stats_div.find('time')
        if time_tag and 'datetime' in time_tag.attrs:
            last_online_str = time_tag['datetime']
            try:
                last_online_time = datetime.fromisoformat(last_online_str.rstrip('Z'))
            except ValueError as e:
                print(f"Ошибка преобразования даты для {admin_url}: {e}")
                return 0
            now = datetime.now()
            time_difference = now - last_online_time
            if time_difference <= timedelta(days=60):
                return 1
            else:
                return 0
        else:
            print(f"Не удалось найти тэг <time> или атрибут 'datetime' для {admin_url}")
            return 0
    except Exception as e:
        print(f"Ошибка при проверке активности администратора {admin_url}: {e}")
        return 0

# Функция для получения списка администраторов клуба
def get_club_admins(club_url):
    try:
        response = requests.get(club_url, headers=headers)
        if response.status_code != 200:
            print(f"Не удалось получить данные для {club_url}")
            return []
        soup = BeautifulSoup(response.text, 'html.parser')
        admin_section = soup.find('section', class_='team-show__meta')
        if not admin_section:
            print(f"Не удалось найти секцию администраторов для {club_url}")
            return []
        admin_links = admin_section.find_all('a', href=True)
        admins = []
        for a in admin_links:
            href = a['href']
            if href.startswith('/@/'):
                admin_url = 'https://lichess.org' + href
                admin_name = href.split('/@/')[1]
                admins.append((admin_url, admin_name))
        return admins
    except Exception as e:
        print(f"Ошибка при получении администраторов для {club_url}: {e}")
        return []

# Основная функция для сбора данных
def collect_data():
    try:
        with open('clubs_domens.txt', 'r') as f:
            club_links = [line.strip() for line in f.readlines()]
        wb = Workbook()
        ws = wb.active
        ws.title = "Клубы и админы"
        ws.append(["Клуб URL", "Админ URL 1", "Имя админа 1", "Статус 1", 
                   "Админ URL 2", "Имя админа 2", "Статус 2", 
                   "Админ URL 3", "Имя админа 3", "Статус 3"])
        for club_url in club_links:
            print(f"Обработка {club_url}...")
            admins = get_club_admins(club_url)
            if not admins:
                print(f"Админы не найдены для {club_url}")
                continue
            row_data = [club_url]
            for admin_url, admin_name in admins:
                activity_status = check_last_online(admin_url)
                row_data.extend([admin_url, admin_name, activity_status])
            ws.append(row_data)
        wb.save("lichess_club_admins.xlsx")
        print("Данные сохранены в 'lichess_club_admins.xlsx'")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

# Функция для вычисления суммы в колонке AF
def calculate_af_column():
    print("Вычисляем значения для колонки AF...")
    filename = 'lichess_club_admins.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    columns_to_sum = [4, 7, 10, 13, 16, 19, 22, 25, 28, 31]  # D, G, J, M, P, S, V, Y, AB, AE
    af_column_index = 32  # AF
    for row in ws.iter_rows(min_row=2, max_col=max(columns_to_sum + [af_column_index]), values_only=False):
        sum_value = sum(cell.value if cell.value is not None else 0 for idx, cell in enumerate(row) if idx + 1 in columns_to_sum)
        row[af_column_index - 1].value = sum_value
    wb.save(filename)
    print("Значения для колонки AF успешно вычислены и записаны.")

if __name__ == "__main__":
    collect_data()
    calculate_af_column()
