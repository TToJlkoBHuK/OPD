import logging
import asyncio
import time
import os
import shutil
from typing import List, Tuple, Dict, Optional # Добавлены типы для аннотаций

from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from aiogram.utils.exceptions import MessageCantBeDeleted
from openpyxl import load_workbook # Для работы с Excel, если используется

# Настройки логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

# --- КОНФИГУРАЦИЯ ---
API_TOKEN = '5425716415:AAF1K7oads37BfNpvCTpBqDQOBcGzGnK0Ww' # ВАШ ТОКЕН
ADMIN_IDS = [1881684121, 5312321185] # ID администраторов

# --- ФАЙЛЫ ДАННЫХ ---
PROGRESS_FILE = "user_progress.txt" # Прогресс получения групп пользователями
BANS_FILE = "user_bans.txt"         # Информация о банах пользователей
USERS_FILE = "users.txt"            # Статусы пользователей (теперь включает статус проверки)
NICKNAMES_FILE = "nicknames.txt"    # Ники пользователей
BROADCAST_TEMPLATE_FILE = "broadcast_template.txt" # Шаблон рассылки (если используется)
GROUPS_DATA_FILE = "groups_data.txt" # Данные о группах из Excel
MEDIA_FOLDER = "user_media"          # Папка для сохранения медиафайлов пользователей

# --- ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ---
bot = Bot(token=API_TOKEN, parse_mode=types.ParseMode.HTML) # Добавлен parse_mode для возможного форматирования
dp = Dispatcher(bot)

groups_data: List[Tuple[str, int]] = []  # Список групп (url, admins_count)
sent_groups: Dict[int, List[int]] = {} # {user_id: [group_index1, group_index2]}
current_group_index: Dict[int, int] = {} # {user_id: last_sent_group_index} - Возможно, не используется активно
user_bans: Dict[int, float] = {}         # {user_id: ban_end_timestamp}
# users_status: Dict[int, str] = {}      # {user_id: status_string} -> ЗАМЕНЕНО НА НОВУЮ СТРУКТУРУ
# Формат: {user_id: {"media_status": "?", "review_status": "?", "admin_status": ""}}
# media_status: Статус медиа пользователя ('?', '✅', '❌') - устанавливается проверяющим
# review_status: Статус проверки пользователем ('?', '✅', '❌') - устанавливается пользователем после проверки
# admin_status: Статус, установленный админом ('', '✅', '❌') - для совместимости и админских пометок
users_status: Dict[int, Dict[str, str]] = {}

user_nicknames: Dict[int, str] = {}    # {user_id: nickname}
media_groups: Dict[int, List[types.Message]] = {} # Временное хранение медиа для группировки (возможно, устарело с новым подходом)
last_media_time: Dict[int, float] = {}      # Время последнего получения медиа (возможно, устарело)
message_queue = asyncio.Queue()             # Очередь сообщений для админов
user_media_accumulation: Dict[int, Dict] = {} # {user_id: {"files": [msg1, msg2], "timer_task": task}}

# --- КОНСТАНТЫ ---
SEND_DELAY = 1 # Задержка при отправке сообщений админам
MEDIA_ACCUMULATION_DELAY = 2.0 # Задержка перед обработкой группы медиафайлов (секунды)
BAN_DURATION = 24 * 60 * 60 # Длительность бана в секундах (24 часа)

# --- ИНИЦИАЛИЗАЦИЯ ---
if not os.path.exists(MEDIA_FOLDER):
    os.makedirs(MEDIA_FOLDER)

# --- ФУНКЦИИ ЗАГРУЗКИ/СОХРАНЕНИЯ ДАННЫХ ---

def load_groups_data() -> Optional[List[Tuple[str, int]]]:
    """Загружает данные о группах из файла."""
    if not os.path.exists(GROUPS_DATA_FILE):
        return None
    groups = []
    try:
        with open(GROUPS_DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 2:
                    try:
                        groups.append((parts[0], int(parts[1])))
                    except ValueError:
                        logging.warning(f"Некорректное число админов в строке файла групп: {line.strip()}")
        return groups
    except Exception as e:
        logging.error(f"Ошибка при загрузке данных групп: {e}")
        return None

def save_groups_data(groups: List[Tuple[str, int]]):
    """Сохраняет данные о группах в файл."""
    try:
        with open(GROUPS_DATA_FILE, "w", encoding="utf-8") as f:
            for group in groups:
                f.write(f"{group[0]},{group[1]}\n")
    except Exception as e:
        logging.error(f"Ошибка при сохранении данных групп: {e}")

def load_data():
    """Загружает все данные бота из файлов."""
    global groups_data, sent_groups, current_group_index, user_bans, users_status, user_nicknames

    groups_data = load_groups_data() or []

    # Загрузка прогресса пользователей
    sent_groups = {}
    current_group_index = {} # Сбрасываем, т.к. он не сохраняется/загружается явно
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(",")
                    if len(parts) >= 1:
                        try:
                            user_id = int(parts[0])
                            valid_indices = [int(index) for index in parts[1:] if index.isdigit()]
                            sent_groups[user_id] = valid_indices
                            # current_group_index не восстанавливаем, т.к. логика его использования не ясна
                        except ValueError:
                            logging.warning(f"Ошибка парсинга строки прогресса: {line.strip()}")
        except Exception as e:
            logging.error(f"Ошибка при загрузке файла прогресса ({PROGRESS_FILE}): {e}")

    # Загрузка банов
    user_bans = {}
    if os.path.exists(BANS_FILE):
        try:
            with open(BANS_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(",")
                    if len(parts) == 2:
                        try:
                            user_bans[int(parts[0])] = float(parts[1])
                        except ValueError:
                            logging.warning(f"Ошибка парсинга строки бана: {line.strip()}")
        except Exception as e:
            logging.error(f"Ошибка при загрузке файла банов ({BANS_FILE}): {e}")

    # Загрузка статусов пользователей (НОВЫЙ ФОРМАТ)
    users_status = {}
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(",", 3) # Ожидаем user_id, media_status, review_status, admin_status
                    if len(parts) == 4:
                        try:
                            user_id = int(parts[0])
                            users_status[user_id] = {
                                "media_status": parts[1],
                                "review_status": parts[2],
                                "admin_status": parts[3]
                            }
                        except ValueError:
                            logging.warning(f"Ошибка парсинга user_id в строке статуса: {line.strip()}")
                    # Обработка старого формата для совместимости (если нужно)
                    elif len(parts) == 2:
                         try:
                            user_id = int(parts[0])
                            # Инициализируем новым форматом, сохраняя старый админский статус
                            users_status[user_id] = {
                                "media_status": "?",
                                "review_status": "?",
                                "admin_status": parts[1] # Старый статус идет в admin_status
                            }
                            logging.info(f"Пользователь {user_id} переведен на новый формат статуса.")
                         except ValueError:
                             logging.warning(f"Ошибка парсинга user_id в строке статуса (старый формат): {line.strip()}")
        except Exception as e:
            logging.error(f"Ошибка при загрузке файла статусов ({USERS_FILE}): {e}")

    # Загрузка никнеймов
    user_nicknames = {}
    if os.path.exists(NICKNAMES_FILE):
        try:
            with open(NICKNAMES_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(":", 1) # Используем split с ограничением 1
                    if len(parts) == 2:
                        try:
                            user_nicknames[int(parts[0])] = parts[1]
                        except ValueError:
                            logging.warning(f"Ошибка парсинга user_id в строке никнейма: {line.strip()}")
        except Exception as e:
            logging.error(f"Ошибка при загрузке файла никнеймов ({NICKNAMES_FILE}): {e}")

    logging.info("Данные успешно загружены.")

def save_data():
    """Сохраняет все данные бота в файлы."""
    # Сохранение прогресса
    try:
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            for user_id, group_indices in sent_groups.items():
                f.write(f"{user_id},{','.join(map(str, group_indices))}\n")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла прогресса ({PROGRESS_FILE}): {e}")

    # Сохранение банов
    try:
        with open(BANS_FILE, "w", encoding="utf-8") as f:
            # Сохраняем только активные баны
            active_bans = {uid: t for uid, t in user_bans.items() if time.time() < t}
            for user_id, ban_time in active_bans.items():
                f.write(f"{user_id},{ban_time}\n")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла банов ({BANS_FILE}): {e}")

    # Сохранение статусов (НОВЫЙ ФОРМАТ)
    try:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            for user_id, status_dict in users_status.items():
                f.write(f"{user_id},{status_dict['media_status']},{status_dict['review_status']},{status_dict['admin_status']}\n")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла статусов ({USERS_FILE}): {e}")

    # Сохранение никнеймов
    try:
        with open(NICKNAMES_FILE, "w", encoding="utf-8") as f:
            for user_id, nickname in user_nicknames.items():
                f.write(f"{user_id}:{nickname}\n")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла никнеймов ({NICKNAMES_FILE}): {e}")

    # logging.info("Данные сохранены.") # Можно раскомментировать для отладки, но может быть слишком часто

def load_broadcast_template() -> str:
    """Загружает шаблон рассылки из файла. Возвращает текст шаблона или сообщение об ошибке."""
    default_template = "Шаблон рассылки не найден. Пожалуйста, настройте его."
    try:
        with open(BROADCAST_TEMPLATE_FILE, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        logging.warning("Файл шаблона рассылки не найден. Возвращен шаблон по умолчанию.")
        return default_template
    except Exception as e:
        logging.error(f"Ошибка при чтении файла шаблона: {e}")
        return f"Ошибка загрузки шаблона: {str(e)}"

# --- ФУНКЦИИ ПРОВЕРКИ И УПРАВЛЕНИЯ ПОЛЬЗОВАТЕЛЯМИ ---

def is_user_banned(user_id: int) -> bool:
    """Проверяет, забанен ли пользователь. Удаляет истекшие баны."""
    if user_id in user_bans:
        if time.time() < user_bans[user_id]:
            return True
        else:
            # Бан истек, удаляем его и сбрасываем статус (если он был связан с баном)
            logging.info(f"Бан для пользователя {user_id} истек.")
            del user_bans[user_id]
            # Сбрасываем только admin_status, оставляя статусы проверки
            if user_id in users_status:
                 users_status[user_id]["admin_status"] = "" # Сбрасываем только админский статус
            save_data()
            return False
    return False

def get_active_user_ids() -> List[int]:
    """Возвращает отсортированный список ID активных (не админ, не забанен) пользователей."""
    active_users = []
    for user_id in users_status:
        if user_id not in ADMIN_IDS and not is_user_banned(user_id):
            active_users.append(user_id)
    active_users.sort() # Сортируем для стабильного порядка
    return active_users

def get_review_target(reviewer_id: int) -> Optional[int]:
    """Определяет ID пользователя, чье медиа должен проверить reviewer_id (циклический сдвиг)."""
    active_users = get_active_user_ids()
    if not active_users:
        logging.warning("Нет активных пользователей для назначения проверки.")
        return None
    if reviewer_id not in active_users:
        logging.warning(f"Пользователь {reviewer_id} не является активным и не может проверять.")
        return None # Сам ревьюер не активен

    try:
        reviewer_index = active_users.index(reviewer_id)
    except ValueError:
        # Этого не должно произойти, если get_active_user_ids работает верно
        logging.error(f"Не удалось найти активного пользователя {reviewer_id} в списке активных.")
        return None

    if len(active_users) == 1:
        logging.info(f"Только один активный пользователь {reviewer_id}. Некого проверять.")
        return None # Некого проверять

    target_index = (reviewer_index + 1) % len(active_users)
    target_user_id = active_users[target_index]
    logging.info(f"Пользователю {reviewer_id} назначен для проверки пользователь {target_user_id}.")
    return target_user_id

def update_review_status(reviewer_id: int, target_user_id: int, is_approved: bool):
    """Обновляет статусы проверки для ревьюера и цели."""
    review_symbol = "✅" if is_approved else "❌"

    # Обновляем статус ревьюера (он выполнил проверку)
    if reviewer_id in users_status:
        users_status[reviewer_id]["review_status"] = review_symbol
        logging.info(f"Статус проверки для {reviewer_id} обновлен на {review_symbol}.")
    else:
        logging.warning(f"Не найден статус для ревьюера {reviewer_id} при обновлении статуса проверки.")

    # Обновляем статус цели (его медиа проверили)
    if target_user_id in users_status:
        users_status[target_user_id]["media_status"] = review_symbol
        logging.info(f"Статус медиа для {target_user_id} обновлен на {review_symbol}.")
    else:
        logging.warning(f"Не найден статус для цели {target_user_id} при обновлении статуса медиа.")

    save_data()

# --- КЛАВИАТУРЫ ---

def get_user_keyboard() -> InlineKeyboardMarkup:
    """Возвращает клавиатуру для обычного пользователя."""
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("Получить группу 🔗", callback_data="get_next_group"),
        InlineKeyboardButton("Проверить медиа 🧐", callback_data="check_media") # Новая кнопка
    )
    keyboard.add(
         InlineKeyboardButton("Я получил бан 😥", callback_data="ban_user")
    )
    keyboard.add(InlineKeyboardButton("Шаблон рассылки", callback_data="show_broadcast_template"))
    keyboard.add(InlineKeyboardButton("Инструкция", callback_data="send_instruction"))
    return keyboard

def get_admin_panel() -> ReplyKeyboardMarkup:
    """Возвращает клавиатуру с основными командами администратора."""
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(KeyboardButton("/run"), KeyboardButton("/send_groups")) # TODO: Проверить назначение /run и /send_groups
    keyboard.add(KeyboardButton("/users"), KeyboardButton("Изменить шаблон")) # TODO: Проверить назначение "Изменить шаблон"
    # Можно добавить /reset_groups, если нужно
    keyboard.add(KeyboardButton("/reset_groups"))
    return keyboard

def get_users_list_keyboard() -> InlineKeyboardMarkup:
    """Возвращает клавиатуру со списком пользователей для админа."""
    keyboard = InlineKeyboardMarkup(row_width=1)
    if not users_status:
         keyboard.add(InlineKeyboardButton("Нет зарегистрированных пользователей", callback_data="no_users"))
         return keyboard

    sorted_user_ids = sorted(users_status.keys()) # Сортируем для единообразия

    for user_id in sorted_user_ids:
        if user_id in ADMIN_IDS: continue # Пропускаем админов

        status_info = users_status.get(user_id, {"media_status": "?", "review_status": "?", "admin_status": ""})
        media_s = status_info['media_status']
        review_s = status_info['review_status']
        admin_s = status_info['admin_status'] # Админская пометка

        # Формируем строку статуса: Медиа | Ревью | Админ (если есть)
        status_str = f"{media_s} | {review_s}"
        if admin_s: # Добавляем админский статус, если он не пустой
            status_str += f" | {admin_s}"

        # Получаем никнейм или ID
        nickname = user_nicknames.get(user_id)
        if nickname:
             button_text = f"@{nickname} | Статус: {status_str}"
        else:
             button_text = f"ID: {user_id} | Статус: {status_str}"

        # Добавляем кнопку бана, если пользователь забанен
        if is_user_banned(user_id):
             button_text += " (🚫 Забанен)"

        keyboard.add(InlineKeyboardButton(button_text, callback_data=f"user_select:{user_id}"))

    return keyboard

def get_admin_user_control_keyboard(user_id: int) -> InlineKeyboardMarkup:
    """Возвращает клавиатуру для управления конкретным пользователем (админ)."""
    keyboard = InlineKeyboardMarkup(row_width=3)
    # Кнопки для админского статуса (третий символ в статусе)
    keyboard.add(
        InlineKeyboardButton(text="Админ: ❌", callback_data=f"set_admin_status:{user_id}:❌"),
        InlineKeyboardButton(text="Админ: ✅", callback_data=f"set_admin_status:{user_id}:✅"),
        InlineKeyboardButton(text="Админ: Очистить", callback_data=f"set_admin_status:{user_id}:")
    )
    # Кнопка снятия бана (появляется, только если юзер забанен)
    if is_user_banned(user_id):
        keyboard.add(InlineKeyboardButton(text="Снять бан 🟢", callback_data=f"remove_ban:{user_id}"))
    else:
        # Можно добавить кнопку бана, если нужно
        # keyboard.add(InlineKeyboardButton(text="Забанить 🔴", callback_data=f"admin_ban:{user_id}")) # Пример
        pass # Пока не добавляем кнопку бана здесь

    keyboard.add(InlineKeyboardButton(text="Просмотреть медиа 🖼️", callback_data=f"view_media:{user_id}"))
    keyboard.add(InlineKeyboardButton(text="« Назад к списку", callback_data="return_to_users_list"))
    return keyboard

def get_review_keyboard(target_user_id: int) -> InlineKeyboardMarkup:
    """Возвращает клавиатуру для оценки медиа другого пользователя."""
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("✅ Одобрить", callback_data=f"review_action:{target_user_id}:approve"),
        InlineKeyboardButton("❌ Отклонить", callback_data=f"review_action:{target_user_id}:reject")
    )
    keyboard.add(InlineKeyboardButton("« Назад в меню", callback_data="return_to_menu"))
    return keyboard

# --- ОБРАБОТКА EXCEL И ГРУПП ---
# Эти функции оставлены как есть, т.к. сказано не менять их

def process_excel() -> List[Tuple[str, int]]:
    logging.info("Обработка Excel и формирование групп...")
    # Убедитесь, что файл существует
    filename = 'lichess_club_admins.xlsx'
    if not os.path.exists(filename):
        logging.error(f"Файл {filename} не найден!")
        return []

    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        # Индекс колонки 'AF'. Нумерация в openpyxl начинается с 1.
        # A=1, B=2, ..., Z=26, AA=27, ..., AE=31, AF=32
        af_column_index = 32
        data = []
        # Пропускаем заголовок (min_row=2)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Проверяем, что в строке достаточно столбцов
            if len(row) >= af_column_index:
                club_url = row[0] # Первый столбец (индекс 0)
                active_admins_raw = row[af_column_index - 1] # Значение в колонке AF

                # Пытаемся преобразовать количество админов в число, обрабатываем ошибки
                try:
                    active_admins = int(active_admins_raw) if active_admins_raw is not None else 0
                except (ValueError, TypeError):
                    logging.warning(f"Не удалось прочитать число админов для {club_url} (значение: {active_admins_raw}). Установлено в 0.")
                    active_admins = 0

                if club_url: # Добавляем только если есть URL клуба
                    data.append((str(club_url), active_admins))
                else:
                    logging.warning("Пропущена строка без URL клуба.")
            else:
                logging.warning(f"Пропущена строка {row}, т.к. в ней меньше {af_column_index} столбцов.")

        # Убираем дубликаты (если нужны) и сортируем
        unique_data = list(set(data))
        unique_data_sorted = sorted(unique_data, key=lambda x: x[1], reverse=True)
        save_groups_data(unique_data_sorted) # Сохраняем группы в файл
        logging.info(f"Обработка Excel завершена. Найдено {len(unique_data_sorted)} уникальных записей.")
        return unique_data_sorted
    except Exception as e:
        logging.exception(f"Критическая ошибка при обработке Excel файла {filename}: {e}")
        return []

# Функция create_group не используется напрямую в текущей логике бота, но оставлена
def create_group(data_sorted, group_size, target_sum):
    current_group = []
    current_sum = 0
    removed_indices = []
    for i, (club_url, admins) in enumerate(data_sorted):
        if len(current_group) < group_size and current_sum + admins <= target_sum:
            current_group.append((club_url, admins))
            current_sum += admins
            removed_indices.append(i)
        if len(current_group) == group_size or current_sum >= target_sum:
            break
    # Создаем новый список без удаленных элементов
    new_data_sorted = [item for i, item in enumerate(data_sorted) if i not in removed_indices]
    return current_group, new_data_sorted

# --- ФОНОВЫЕ ЗАДАЧИ ---

async def process_message_queue():
    """Обрабатывает очередь сообщений для отправки админам."""
    while True:
        try:
            chat_id, message_text, media = await message_queue.get()
            try:
                if media:
                    # Отправляем медиагруппы порциями по 10
                    for i in range(0, len(media), 10):
                        media_chunk = media[i:i + 10]
                        await bot.send_media_group(chat_id=chat_id, media=media_chunk)
                        await asyncio.sleep(0.5) # Небольшая пауза между пачками
                    logging.info(f"Медиагруппа ({len(media)} шт.) отправлена админу {chat_id}.")
                elif message_text:
                    await bot.send_message(chat_id=chat_id, text=message_text)
                    logging.info(f"Текстовое сообщение отправлено админу {chat_id}.")
            except Exception as e:
                logging.error(f"Ошибка при отправке сообщения/медиа админу {chat_id}: {e}")
            finally:
                await asyncio.sleep(SEND_DELAY) # Задержка перед следующим сообщением
                message_queue.task_done()
        except asyncio.CancelledError:
            logging.info("Очередь обработки сообщений остановлена.")
            break
        except Exception as e:
            logging.error(f"Критическая ошибка в process_message_queue: {e}")
            await asyncio.sleep(5) # Пауза перед повторной попыткой

async def restore_ban_if_inactive(user_id: int):
    """Задача для автоматического бана неактивного пользователя через 24 часа после разбана."""
    await asyncio.sleep(BAN_DURATION) # Ждем 24 часа
    if user_id in users_status and not is_user_banned(user_id):
        # Проверяем активность - здесь нужна логика определения активности
        # Например, по времени последнего сообщения или нажатия кнопки.
        # Пока просто баним, если не забанен снова вручную.
        logging.info(f"Пользователь {user_id} не проявил активность после разбана. Автоматический бан.")
        user_bans[user_id] = time.time() + BAN_DURATION
        if user_id in users_status:
            users_status[user_id]["admin_status"] = "" # Сбрасываем админский статус при автобане
        save_data()
        try:
            await bot.send_message(
                user_id,
                "Вы были автоматически заблокированы за неактивность после снятия предыдущего бана."
            )
        except Exception as e:
            logging.warning(f"Не удалось отправить сообщение об автобане пользователю {user_id}: {e}")

# --- ОБРАБОТЧИКИ СООБЩЕНИЙ И КОМАНД ---

@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username

    # Обновляем никнейм при старте
    if username and user_nicknames.get(user_id) != username:
        user_nicknames[user_id] = username
        logging.info(f"Обновлен никнейм для {user_id} на @{username} при старте.")
        save_data() # Сохраняем сразу
    elif user_id not in user_nicknames: # Если ника нет и не было
         user_nicknames[user_id] = "" # Сохраняем пустую строку как маркер отсутствия ника
         save_data()

    # Добавляем пользователя в users_status, если его там нет (НОВЫЙ ФОРМАТ)
    if user_id not in users_status:
        is_new_user = True
        users_status[user_id] = {"media_status": "?", "review_status": "?", "admin_status": ""}
        logging.info(f"Новый пользователь {user_id} (@{username or 'ID'}) зарегистрирован со статусом по умолчанию.")
        save_data()
    else:
        is_new_user = False

    if user_id in ADMIN_IDS:
        await message.reply(
            "Привет, администратор! 👋 Вот ваша панель управления:",
            reply_markup=get_admin_panel()
        )
    else:
        # Проверяем бан ПОСЛЕ регистрации статуса
        if is_user_banned(user_id):
            ban_end_time = user_bans.get(user_id, time.time())
            remaining_time = ban_end_time - time.time()
            hours, rem = divmod(remaining_time, 3600)
            minutes, _ = divmod(rem, 60)
            time_left_str = f"{int(hours)} ч {int(minutes)} мин" if hours > 0 else f"{int(minutes)} мин"
            await message.reply(f"Вы заблокированы. 🚫\nОсталось примерно: {time_left_str}")
            return

        keyboard = get_user_keyboard()
        welcome_message = "Привет! 👋 Нажмите кнопку ниже, чтобы получить группу или проверить медиа."
        if is_new_user:
            welcome_message = "Добро пожаловать! 👋\n" + welcome_message
        await message.reply(welcome_message, reply_markup=keyboard)

# Команды администратора /run и /send_groups (оставлены без изменений)
# TODO: Реализовать или проверить функционал этих команд, если он нужен.
@dp.message_handler(commands=['run'], user_id=ADMIN_IDS)
async def run_command(message: types.Message):
    await message.reply("Команда /run выполняется...")
    # Добавьте сюда логику команды /run

@dp.message_handler(commands=['send_groups'], user_id=ADMIN_IDS)
async def send_groups_command(message: types.Message):
    await message.reply("Команда /send_groups выполняется...")
    # Добавьте сюда логику команды /send_groups

@dp.message_handler(commands=['reset_groups'], user_id=ADMIN_IDS)
async def reset_groups(message: types.Message):
    global sent_groups, groups_data
    logging.warning(f"Администратор {message.from_user.id} инициировал сброс прогресса и пересоздание групп.")
    sent_groups = {}
    # Пересоздаем группы из Excel
    groups_data = process_excel()
    # Сбрасываем прогресс у всех пользователей (удаляем файл прогресса или очищаем словарь)
    sent_groups.clear()
    # Сбрасываем статусы проверки медиа у всех пользователей
    for user_id in users_status:
        users_status[user_id]["media_status"] = "?"
        users_status[user_id]["review_status"] = "?"
        # admin_status не трогаем
    save_data() # Сохраняем пустой прогресс и сброшенные статусы
    await message.answer("✅ Прогресс всех пользователей сброшен.\n✅ Статусы проверки медиа сброшены.\n✅ Группы пересозданы из Excel.")

@dp.message_handler(commands=['users'], user_id=ADMIN_IDS)
async def list_users(message: types.Message):
    if not users_status:
        await message.answer("Нет зарегистрированных пользователей.")
        return
    await message.answer("👥 Список пользователей:", reply_markup=get_users_list_keyboard())

admin_states = {} # {user_id: state} - Для машины состояний админа (например, ожидание шаблона)

# Обработчик кнопки "Изменить шаблон" (оставлен без изменений)
@dp.message_handler(lambda message: message.text == "Изменить шаблон", user_id=ADMIN_IDS)
async def request_broadcast_template(message: types.Message):
     user_id = message.from_user.id
     admin_states[user_id] = "waiting_for_template"
     try:
         with open(BROADCAST_TEMPLATE_FILE, "r", encoding="utf-8") as f:
              current_template = f.read()
         await message.answer(f"Текущий шаблон:\n```\n{current_template}\n```\nОтправьте новый текст шаблона.", parse_mode="Markdown")
     except FileNotFoundError:
         await message.answer("Файл шаблона не найден. Отправьте текст для нового шаблона.")
     except Exception as e:
         logging.error(f"Ошибка чтения файла шаблона: {e}")
         await message.answer("Ошибка при чтении текущего шаблона. Отправьте текст для нового шаблона.")

# Функция сохранения шаблона (вынесена для ясности)
def save_broadcast_template(template_text: str):
    try:
        with open(BROADCAST_TEMPLATE_FILE, "w", encoding="utf-8") as f:
            f.write(template_text)
        logging.info("Шаблон рассылки обновлен.")
    except Exception as e:
        logging.error(f"Ошибка записи файла шаблона: {e}")

# Обработка нового шаблона (оставлен без изменений)
@dp.message_handler(lambda message: message.from_user.id in admin_states and admin_states[message.from_user.id] == "waiting_for_template", user_id=ADMIN_IDS)
async def update_broadcast_template(message: types.Message):
    user_id = message.from_user.id
    new_template = message.text
    save_broadcast_template(new_template)
    if user_id in admin_states:
        del admin_states[user_id] # Выходим из состояния ожидания
    await message.answer("✅ Шаблон успешно обновлен!")

# --- ОБРАБОТЧИКИ КОЛЛБЭКОВ ---

@dp.callback_query_handler(lambda c: c.data == "return_to_menu")
async def return_to_menu(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    if is_user_banned(user_id):
        await callback_query.answer("Вы заблокированы.", show_alert=True)
        # Можно удалить сообщение с кнопками, если нужно
        # try: await bot.delete_message(user_id, callback_query.message.message_id)
        # except: pass
        return

    keyboard = get_user_keyboard()
    try:
        # Редактируем сообщение, чтобы убрать старую клавиатуру и показать новую
        await bot.edit_message_text("Вы в главном меню.", user_id, callback_query.message.message_id, reply_markup=keyboard)
        await callback_query.answer() # Отвечаем на коллбэк
    except Exception as e:
        logging.warning(f"Не удалось отредактировать сообщение для возврата в меню пользователя {user_id}: {e}. Отправляем новое.")
        # Если редактирование не удалось (например, сообщение слишком старое), отправляем новое
        await bot.send_message(user_id, "Вы в главном меню.", reply_markup=keyboard)
        # Пытаемся удалить старое сообщение, если оно еще доступно
        try: await bot.delete_message(user_id, callback_query.message.message_id)
        except: pass
        await callback_query.answer()

@dp.callback_query_handler(lambda c: c.data == "return_to_users_list", user_id=ADMIN_IDS)
async def return_to_users_list(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    try:
        # Редактируем сообщение, чтобы показать список пользователей
        await bot.edit_message_text(
            chat_id=admin_id,
            message_id=callback_query.message.message_id,
            text="👥 Список пользователей:",
            reply_markup=get_users_list_keyboard()
        )
        await callback_query.answer()
    except Exception as e:
        logging.warning(f"Не удалось отредактировать сообщение для возврата к списку пользователей: {e}")
        await callback_query.answer("Не удалось обновить список.", show_alert=True)
        # Можно попробовать отправить новый список
        await bot.send_message(admin_id, "👥 Список пользователей:", reply_markup=get_users_list_keyboard())


@dp.callback_query_handler(lambda c: c.data.startswith("remove_ban:"), user_id=ADMIN_IDS)
async def remove_ban(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    try:
        _, user_id_str = callback_query.data.split(":")
        user_id = int(user_id_str)
    except ValueError:
        await callback_query.answer("Ошибка ID пользователя.", show_alert=True)
        return

    if user_id not in user_bans:
        await callback_query.answer("Пользователь не заблокирован.", show_alert=False)
        return

    # Снимаем бан
    del user_bans[user_id]
    # Сбрасываем админский статус при снятии бана админом
    if user_id in users_status:
         users_status[user_id]["admin_status"] = ""
    save_data()
    logging.info(f"Администратор {admin_id} снял бан с пользователя {user_id}.")

    # Уведомляем пользователя
    try:
        await bot.send_message(
            user_id,
            "Администратор снял вашу блокировку. ✅\nВы можете продолжить работу.",
            reply_markup=get_user_keyboard()
        )
    except Exception as e:
        logging.warning(f"Не удалось отправить уведомление о снятии бана пользователю {user_id}: {e}")
        await callback_query.answer("Бан снят, но не удалось уведомить пользователя.", show_alert=True)

    # Обновляем сообщение админа
    try:
        await bot.edit_message_text(
            chat_id=admin_id,
            message_id=callback_query.message.message_id,
            text=f"Блокировка пользователя ID: {user_id} успешно снята.",
            reply_markup=get_admin_user_control_keyboard(user_id) # Показываем обновленную клавиатуру управления
        )
        await callback_query.answer("Бан снят.")
    except Exception as e:
         logging.warning(f"Не удалось обновить сообщение админа после снятия бана: {e}")
         await callback_query.answer("Бан снят.", show_alert=True) # Сообщаем админу


    # Запускаем таймер автобана за неактивность (если нужно)
    # asyncio.create_task(restore_ban_if_inactive(user_id))


@dp.callback_query_handler(lambda c: c.data.startswith("view_media:"), user_id=ADMIN_IDS)
async def view_media(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    try:
        _, user_id_str = callback_query.data.split(":")
        user_id = int(user_id_str)
    except ValueError:
        await callback_query.answer("Ошибка ID пользователя.", show_alert=True)
        return

    user_folder = os.path.join(MEDIA_FOLDER, str(user_id))
    media_to_send: List[types.InputMedia] = [] # Явно указываем тип

    if os.path.exists(user_folder) and os.path.isdir(user_folder):
        try:
            for file_name in sorted(os.listdir(user_folder)): # Сортируем для порядка
                file_path = os.path.join(user_folder, file_name)
                if not os.path.isfile(file_path): continue # Пропускаем подпапки

                # Определяем тип медиа по расширению
                _, ext = os.path.splitext(file_name)
                ext = ext.lower()

                try:
                    file = types.InputFile(file_path)
                    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                        media_to_send.append(types.InputMediaPhoto(media=file))
                    elif ext in ['.mp4', '.avi', '.mov', '.mkv']:
                         media_to_send.append(types.InputMediaVideo(media=file))
                    elif ext in ['.doc', '.docx', '.pdf', '.txt', '.xls', '.xlsx', '.ppt', '.pptx']:
                         media_to_send.append(types.InputMediaDocument(media=file))
                    # Можно добавить другие типы, например, аудио
                    # elif ext in ['.mp3', '.ogg', '.wav']:
                    #     media_to_send.append(types.InputMediaAudio(media=file))
                    else:
                        logging.info(f"Пропущен неподдерживаемый файл {file_name} для пользователя {user_id}")

                except Exception as e:
                    logging.error(f"Ошибка при подготовке файла {file_path} для отправки: {e}")

        except Exception as e:
            logging.error(f"Ошибка при чтении папки {user_folder}: {e}")
            await callback_query.answer("Ошибка чтения медиа пользователя.", show_alert=True)
            return
    else:
        logging.info(f"Папка медиа для пользователя {user_id} не найдена.")
        # Не выводим alert, просто сообщим в тексте

    # Отправка медиа и сообщения
    try:
        if media_to_send:
            await callback_query.answer("Загружаю медиа...")
            # Удаляем исходное сообщение с кнопкой "Просмотреть медиа"
            await bot.delete_message(admin_id, callback_query.message.message_id)

            # Отправляем группами по 10
            for i in range(0, len(media_to_send), 10):
                chunk = media_to_send[i:i + 10]
                await bot.send_media_group(chat_id=admin_id, media=chunk)
                await asyncio.sleep(0.5) # Пауза

            # Отправляем сообщение с клавиатурой управления после всех медиа
            await bot.send_message(
                chat_id=admin_id,
                text=f"Медиа пользователя ID: {user_id} ({len(media_to_send)} файлов).",
                reply_markup=get_admin_user_control_keyboard(user_id)
            )
        else:
             # Редактируем исходное сообщение, если медиа нет
            await bot.edit_message_text(
                chat_id=admin_id,
                message_id=callback_query.message.message_id,
                text=f"У пользователя ID: {user_id} нет сохраненных медиа.",
                reply_markup=get_admin_user_control_keyboard(user_id)
            )
            await callback_query.answer("Медиа не найдено.")

    except Exception as e:
        logging.error(f"Ошибка при отправке медиа админу {admin_id} для пользователя {user_id}: {e}")
        await callback_query.answer("Ошибка при отправке медиа.", show_alert=True)
        # Возвращаем клавиатуру управления, даже если была ошибка отправки
        try:
             await bot.send_message(admin_id, f"Ошибка отправки медиа для ID: {user_id}.", reply_markup=get_admin_user_control_keyboard(user_id))
        except: pass


@dp.callback_query_handler(lambda c: c.data.startswith("user_select:"), user_id=ADMIN_IDS)
async def select_user(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    try:
        _, user_id_str = callback_query.data.split(":")
        user_id = int(user_id_str)
    except ValueError:
        await callback_query.answer("Ошибка ID пользователя.", show_alert=True)
        return

    if user_id not in users_status:
        await callback_query.answer("Пользователь не найден в системе.", show_alert=True)
        # Обновить список пользователей на случай, если данные устарели
        await bot.edit_message_text(
            chat_id=admin_id,
            message_id=callback_query.message.message_id,
            text="Пользователь не найден. Обновленный список:",
            reply_markup=get_users_list_keyboard()
        )
        return

    # Получаем текущий статус
    status_info = users_status.get(user_id, {"media_status": "?", "review_status": "?", "admin_status": ""})
    media_s = status_info['media_status']
    review_s = status_info['review_status']
    admin_s = status_info['admin_status']
    status_str = f"{media_s} | {review_s}"
    if admin_s: status_str += f" | {admin_s}"
    ban_str = " (🚫 Забанен)" if is_user_banned(user_id) else ""

    # Формируем текст
    nickname = user_nicknames.get(user_id)
    user_display = f"@{nickname}" if nickname else f"ID: {user_id}"
    text = f"Управление пользователем: {user_display}\n" \
           f"Текущий статус (Медиа | Ревью | Админ): {status_str}{ban_str}\n" \
           f"Выберите действие:"

    try:
        await bot.edit_message_text(
            chat_id=admin_id,
            message_id=callback_query.message.message_id,
            text=text,
            reply_markup=get_admin_user_control_keyboard(user_id)
        )
        await callback_query.answer()
    except Exception as e:
        logging.warning(f"Не удалось отредактировать сообщение для выбора пользователя {user_id}: {e}")
        await callback_query.answer("Ошибка отображения управления пользователем.", show_alert=True)


@dp.callback_query_handler(lambda c: c.data.startswith("set_admin_status:"), user_id=ADMIN_IDS)
async def set_user_admin_status(callback_query: types.CallbackQuery):
    admin_id = callback_query.from_user.id
    try:
        _, user_id_str, status = callback_query.data.split(":", 2)
        user_id = int(user_id_str)
    except ValueError:
        await callback_query.answer("Ошибка ID пользователя или статуса.", show_alert=True)
        return

    if user_id not in users_status:
        await callback_query.answer("Пользователь не найден.", show_alert=True)
        return

    # Обновляем только admin_status
    users_status[user_id]["admin_status"] = status
    save_data()
    logging.info(f"Администратор {admin_id} установил админ-статус '{status}' для пользователя {user_id}.")

    # Обновляем сообщение с кнопками управления
    # Получаем актуальный полный статус
    status_info = users_status.get(user_id, {"media_status": "?", "review_status": "?", "admin_status": ""})
    media_s = status_info['media_status']
    review_s = status_info['review_status']
    admin_s = status_info['admin_status']
    status_str = f"{media_s} | {review_s}"
    if admin_s: status_str += f" | {admin_s}"
    ban_str = " (🚫 Забанен)" if is_user_banned(user_id) else ""
    nickname = user_nicknames.get(user_id)
    user_display = f"@{nickname}" if nickname else f"ID: {user_id}"
    text = f"Управление пользователем: {user_display}\n" \
           f"Текущий статус (Медиа | Ревью | Админ): {status_str}{ban_str}\n" \
           f"Статус администратора обновлен. Выберите действие:"

    try:
        await bot.edit_message_text(
            chat_id=admin_id,
            message_id=callback_query.message.message_id,
            text=text,
            reply_markup=get_admin_user_control_keyboard(user_id)
        )
        await callback_query.answer(f"Админ-статус изменен на: {status or 'Пусто'}")
    except Exception as e:
        logging.warning(f"Не удалось обновить сообщение после установки админ-статуса для {user_id}: {e}")
        await callback_query.answer(f"Админ-статус изменен на: {status or 'Пусто'}", show_alert=True)

# --- ОБРАБОТЧИКИ КОЛЛБЭКОВ ПОЛЬЗОВАТЕЛЯ ---

@dp.callback_query_handler(lambda c: c.data == "get_next_group")
async def get_next_group(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    if user_id in ADMIN_IDS:
        await callback_query.answer("Администраторы не получают группы.", show_alert=True)
        return
    if is_user_banned(user_id):
        await callback_query.answer("Вы заблокированы.", show_alert=True)
        return
    if not groups_data:
        await callback_query.answer("Нет доступных групп для выдачи.", show_alert=True)
        return

    # Собираем все индексы групп, которые уже были выданы *всем* пользователям
    used_indices = set()
    for indices in sent_groups.values():
        used_indices.update(indices)

    next_index = -1
    # Ищем первый неиспользованный индекс
    for i in range(len(groups_data)):
        if i not in used_indices:
            next_index = i
            break

    if next_index == -1:
        await callback_query.answer("Все группы уже были розданы.", show_alert=True)
        return

    # Добавляем индекс к списку полученных пользователем
    if user_id not in sent_groups:
        sent_groups[user_id] = []
    sent_groups[user_id].append(next_index)
    save_data() # Сохраняем прогресс

    # Получаем данные группы
    club_url, active_admins = groups_data[next_index]
    group_number = next_index + 1
    total_groups = len(groups_data)

    group_message = f"✅ Ваша группа #{group_number} (из {total_groups}):\n\n" \
                    f"🔗 Ссылка: {club_url}\n" \
                    f"👥 Активных админов: {active_admins}"

    # Отправляем сообщение с группой и основной клавиатурой пользователя
    try:
        # Сначала отвечаем на коллбэк, чтобы кнопка не висела в состоянии загрузки
        await callback_query.answer()
        # Удаляем предыдущее сообщение с кнопками
        try:
            await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)
        except MessageCantBeDeleted:
             logging.warning(f"Не удалось удалить старое сообщение для {user_id}")
        except Exception as e:
             logging.warning(f"Ошибка при удалении старого сообщения для {user_id}: {e}")

        # Отправляем новое сообщение
        await bot.send_message(user_id, group_message, reply_markup=get_user_keyboard(), disable_web_page_preview=True)
        logging.info(f"Пользователь {user_id} получил группу #{group_number}: {club_url}")

    except Exception as e:
        logging.error(f"Ошибка при отправке группы пользователю {user_id}: {e}")
        # Попытка отката прогресса, если отправка не удалась
        if next_index in sent_groups.get(user_id, []):
            sent_groups[user_id].remove(next_index)
            save_data()
            logging.warning(f"Откат прогресса для пользователя {user_id} после ошибки отправки группы {next_index}.")
        await bot.send_message(user_id, "Произошла ошибка при получении группы. Попробуйте позже.", reply_markup=get_user_keyboard())


@dp.callback_query_handler(lambda c: c.data == "ban_user")
async def ban_user(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    if user_id in ADMIN_IDS:
         await callback_query.answer("Администратор не может забанить себя.", show_alert=True)
         return
    if user_id not in users_status:
        await callback_query.answer("Вы не зарегистрированы в системе.", show_alert=True)
        return
    if is_user_banned(user_id):
        await callback_query.answer("Вы уже заблокированы.", show_alert=False)
        return

    # Баним пользователя
    ban_end_time = time.time() + BAN_DURATION
    user_bans[user_id] = ban_end_time
    # Сбрасываем админский статус при самобане
    if user_id in users_status:
        users_status[user_id]["admin_status"] = ""
    save_data()
    logging.info(f"Пользователь {user_id} инициировал самобан.")

    # Отправляем сообщение о бане
    try:
        # Отвечаем на коллбэк
        await callback_query.answer("Вы заблокированы на 24 часа.")
        # Удаляем сообщение с кнопками
        try:
            await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)
        except Exception as e:
             logging.warning(f"Не удалось удалить сообщение перед отправкой бана {user_id}: {e}")

        await bot.send_message(user_id, "Вы запросили блокировку и были заблокированы на 24 часа. 🚫")
    except Exception as e:
        logging.error(f"Ошибка при отправке сообщения о бане пользователю {user_id}: {e}")


@dp.callback_query_handler(lambda c: c.data == "check_media")
async def handle_check_media(callback_query: types.CallbackQuery):
    """Обработчик кнопки 'Проверить медиа'."""
    reviewer_id = callback_query.from_user.id
    if is_user_banned(reviewer_id):
        await callback_query.answer("Вы заблокированы и не можете проверять медиа.", show_alert=True)
        return

    # Определяем, кого должен проверять пользователь
    target_user_id = get_review_target(reviewer_id)

    if target_user_id is None:
        await callback_query.answer("Не удалось найти пользователя для проверки.\nВозможно, вы единственный активный пользователь.", show_alert=True)
        return

    target_user_folder = os.path.join(MEDIA_FOLDER, str(target_user_id))
    media_to_send: List[types.InputMedia] = []

    # Проверяем наличие папки и собираем медиа
    if os.path.exists(target_user_folder) and os.path.isdir(target_user_folder):
        try:
            for file_name in sorted(os.listdir(target_user_folder)):
                file_path = os.path.join(target_user_folder, file_name)
                if not os.path.isfile(file_path): continue

                _, ext = os.path.splitext(file_name)
                ext = ext.lower()

                try:
                    file = types.InputFile(file_path)
                    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                        media_to_send.append(types.InputMediaPhoto(media=file))
                    elif ext in ['.mp4', '.avi', '.mov', '.mkv']:
                         media_to_send.append(types.InputMediaVideo(media=file))
                    elif ext in ['.doc', '.docx', '.pdf', '.txt', '.xls', '.xlsx', '.ppt', '.pptx']:
                         media_to_send.append(types.InputMediaDocument(media=file))
                    else:
                        logging.info(f"Пропущен неподдерживаемый файл {file_name} при проверке для {reviewer_id} (файл от {target_user_id})")

                    # Ограничение на общее количество файлов для показа (можно убрать или изменить)
                    # if len(media_to_send) >= 50: # Например, показывать не более 50 файлов
                    #     logging.warning(f"Достигнут лимит показа медиа ({len(media_to_send)}) для проверки пользователем {reviewer_id} (медиа от {target_user_id})")
                    #     break

                except Exception as e:
                    logging.error(f"Ошибка при подготовке файла {file_path} для проверки: {e}")

        except Exception as e:
            logging.error(f"Ошибка при чтении папки {target_user_folder} для проверки: {e}")
            await callback_query.answer("Ошибка чтения медиа пользователя для проверки.", show_alert=True)
            return
    else:
        logging.info(f"Папка медиа для пользователя {target_user_id} (цель проверки для {reviewer_id}) не найдена.")

    # Отправляем медиа и кнопки оценки
    try:
        # Отвечаем на коллбэк
        await callback_query.answer("Загружаю медиа для проверки...")
        # Удаляем старое сообщение с кнопками
        try: await bot.delete_message(reviewer_id, callback_query.message.message_id)
        except: pass

        target_nickname = user_nicknames.get(target_user_id)
        target_display = f"@{target_nickname}" if target_nickname else f"ID: {target_user_id}"

        if media_to_send:
            await bot.send_message(reviewer_id, f"Пожалуйста, проверьте медиа от пользователя {target_display}:")
            # Отправляем группами по 10
            for i in range(0, len(media_to_send), 10):
                chunk = media_to_send[i:i + 10]
                await bot.send_media_group(chat_id=reviewer_id, media=chunk)
                await asyncio.sleep(0.5) # Пауза

            # Отправляем сообщение с кнопками оценки
            await bot.send_message(
                chat_id=reviewer_id,
                text=f"Оцените медиа ({len(media_to_send)} файлов) от {target_display}:",
                reply_markup=get_review_keyboard(target_user_id)
            )
        else:
            # Если медиа нет, сообщаем и возвращаем обычную клавиатуру
            await bot.send_message(
                chat_id=reviewer_id,
                text=f"У пользователя {target_display} нет медиа для проверки.",
                reply_markup=get_user_keyboard() # Возвращаем в меню
            )

    except Exception as e:
        logging.error(f"Ошибка при отправке медиа для проверки пользователю {reviewer_id}: {e}")
        await bot.send_message(reviewer_id, "Произошла ошибка при загрузке медиа для проверки. Попробуйте позже.", reply_markup=get_user_keyboard())


@dp.callback_query_handler(lambda c: c.data.startswith("review_action:"))
async def handle_review_action(callback_query: types.CallbackQuery):
    """Обработчик нажатия кнопок ✅ или ❌ при проверке медиа."""
    reviewer_id = callback_query.from_user.id
    if is_user_banned(reviewer_id):
        await callback_query.answer("Вы заблокированы.", show_alert=True)
        return

    try:
        _, target_user_id_str, action = callback_query.data.split(":")
        target_user_id = int(target_user_id_str)
        is_approved = (action == "approve")
    except ValueError:
        await callback_query.answer("Ошибка данных.", show_alert=True)
        logging.error(f"Некорректный формат callback_data для review_action: {callback_query.data}")
        return

    # Обновляем статусы
    update_review_status(reviewer_id, target_user_id, is_approved)
    action_text = "одобрили" if is_approved else "отклонили"
    logging.info(f"Пользователь {reviewer_id} {action_text} медиа пользователя {target_user_id}.")

    # Отправляем подтверждение и возвращаем в меню
    try:
        await callback_query.answer(f"Ваш голос ({'✅' if is_approved else '❌'}) учтен.")
        # Удаляем сообщение с кнопками проверки
        try:
            await bot.delete_message(chat_id=reviewer_id, message_id=callback_query.message.message_id)
        except Exception as e:
             logging.warning(f"Не удалось удалить сообщение с кнопками проверки для {reviewer_id}: {e}")

        await bot.send_message(reviewer_id, "Спасибо за проверку! Вы вернулись в главное меню.", reply_markup=get_user_keyboard())

        # Опционально: Уведомить пользователя, чье медиа проверили
        try:
            reviewer_nickname = user_nicknames.get(reviewer_id)
            reviewer_display = f"@{reviewer_nickname}" if reviewer_nickname else f"ID: {reviewer_id}"
            await bot.send_message(
                 target_user_id,
                 f"Пользователь {reviewer_display} проверил ваше медиа и поставил оценку: {'✅' if is_approved else '❌'}"
            )
        except Exception as e:
             logging.warning(f"Не удалось уведомить пользователя {target_user_id} о результате проверки: {e}")


    except Exception as e:
        logging.error(f"Ошибка при обработке результата проверки от {reviewer_id}: {e}")
        await bot.send_message(reviewer_id, "Произошла ошибка при сохранении вашего решения. Попробуйте позже.", reply_markup=get_user_keyboard())

# Обработка кнопки "Шаблон рассылки"
@dp.callback_query_handler(lambda c: c.data == "show_broadcast_template")
async def show_broadcast_template(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    # Загружаем текущий шаблон
    template_text = load_broadcast_template()
    # Клавиатура с кнопкой возврата
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("Вернуться в меню", callback_data="return_to_menu")
    )
    # Отправляем шаблон текста и клавиатуру
    await bot.send_message(user_id, f"```\n{template_text}\n```", reply_markup=keyboard, parse_mode="Markdown")
    # Удаляем старое сообщение с кнопками
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

# Обработка кнопки "Инструкция"
@dp.callback_query_handler(lambda c: c.data == "send_instruction")
async def send_instruction(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    # Путь к файлу инструкции (предполагается, что файл лежит в папке проекта)
    instruction_file = "Инструкция.pdf"

    # Проверяем, существует ли файл
    if not os.path.exists(instruction_file):
        await bot.send_message(user_id, "Извините, инструкция временно недоступна.")
        return

    # Клавиатура с кнопкой возврата
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("Вернуться в меню", callback_data="return_to_menu")
    )

    # Отправляем PDF-файл
    with open(instruction_file, "rb") as file:
        await bot.send_document(
            chat_id=user_id,
            document=file,
            caption="Инструкция по использованию бота и рассылке.",
            reply_markup=keyboard
        )

    # Удаляем старое сообщение с кнопками
    await bot.delete_message(chat_id=user_id, message_id=callback_query.message.message_id)

# --- ОБРАБОТКА МЕДИАФАЙЛОВ ---

async def delayed_media_processing(user_id: int, delay: float):
     """Запускает обработку накопленных медиа с задержкой."""
     await asyncio.sleep(delay)
     logging.debug(f"Таймер обработки медиа для {user_id} истек.")
     # Запускаем основную функцию обработки в фоне
     asyncio.create_task(process_accumulated_media(user_id))

async def process_accumulated_media(user_id: int):
    """Обрабатывает и сохраняет накопленные медиафайлы пользователя."""
    if user_id not in user_media_accumulation:
        logging.debug(f"Нет накопленных медиа для обработки у {user_id}.")
        return

    media_data = user_media_accumulation.pop(user_id, None) # Забираем данные и удаляем из накопления
    if not media_data or not media_data.get("files"):
        logging.debug(f"Нет файлов для обработки у {user_id} после извлечения.")
        return

    messages_to_process = media_data["files"]
    user_folder = os.path.join(MEDIA_FOLDER, str(user_id))
    username = messages_to_process[0].from_user.username or f"ID: {user_id}" # Берем из первого сообщения
    logging.info(f"Начало сохранения {len(messages_to_process)} медиафайлов для {username} ({user_id}).")

    # Очищаем папку пользователя перед сохранением новых файлов
    if os.path.exists(user_folder):
        try:
            shutil.rmtree(user_folder)
            logging.info(f"Папка {user_folder} очищена перед сохранением новых медиа.")
        except Exception as e:
            logging.error(f"Ошибка при очистке папки {user_folder}: {e}")
            # Продолжаем попытку сохранения, но могут быть проблемы

    # Создаем папку (даже если rmtree не сработал, makedirs обработает exist_ok)
    try:
        os.makedirs(user_folder, exist_ok=True)
    except Exception as e:
        logging.error(f"Не удалось создать папку {user_folder}: {e}")
        # Нет смысла продолжать без папки
        # Возможно, стоит уведомить пользователя или админа
        return

    saved_count = 0
    # Используем счетчик для уникальных имен файлов на случай дублирования file_unique_id
    file_counter = 0
    for message in messages_to_process:
        file_info = None
        file_ext = ".dat" # Расширение по умолчанию
        file_name_base = f"{time.time_ns()}_{file_counter}" # Уникальное имя по времени + счетчик

        media_type = "" # Для логирования

        if message.photo:
            file_info = message.photo[-1] # Берем самое большое разрешение
            file_ext = ".jpg"
            # file_name_base = file_info.file_unique_id # Уникальный ID файла не всегда надежен для имени
            media_type = "фото"
        elif message.video:
            file_info = message.video
            file_ext = ".mp4" # Или другое стандартное
            # file_name_base = file_info.file_unique_id
            media_type = "видео"
        elif message.document:
            file_info = message.document
            original_filename = getattr(file_info, 'file_name', '')
            if original_filename:
                 _, ext_from_name = os.path.splitext(original_filename)
                 if ext_from_name:
                     file_ext = ext_from_name.lower()
            # file_name_base = file_info.file_unique_id
            media_type = "документ"
        # Добавить другие типы при необходимости (audio, voice)

        if file_info:
            destination_path = os.path.join(user_folder, f"{file_name_base}{file_ext}")
            try:
                # Скачиваем файл
                await bot.download_file_by_id(file_info.file_id, destination=destination_path)
                saved_count += 1
                file_counter += 1 # Увеличиваем счетчик для следующего файла
                # logging.debug(f"Файл ({media_type}) сохранен: {destination_path}")
            except Exception as e:
                logging.error(f"Ошибка скачивания файла ({media_type}) {file_info.file_id} для {user_id}: {e}")
                # Удаляем частично скачанный файл, если он есть
                if os.path.exists(destination_path):
                    try: os.remove(destination_path)
                    except: pass
        else:
             logging.warning(f"Сообщение {message.message_id} от {user_id} не содержит медиа для сохранения.")

    logging.info(f"Завершено сохранение медиа для {username} ({user_id}). Сохранено {saved_count} из {len(messages_to_process)} файлов в папку {user_folder}.")
    # Сбрасываем статус проверки медиа пользователя на '?', т.к. он отправил новые
    if user_id in users_status:
        users_status[user_id]["media_status"] = "?"
        save_data()
        logging.info(f"Статус медиа для {user_id} сброшен на '?' после загрузки новых файлов.")
    # Опционально: Уведомить пользователя об успешном сохранении
    # try:
    #      await bot.send_message(user_id, f"✅ Ваши {saved_count} медиафайлов успешно загружены и ожидают проверки.")
    # except Exception as e:
    #      logging.warning(f"Не удалось уведомить {user_id} о сохранении медиа: {e}")

# --- ОБЩИЙ ОБРАБОТЧИК СООБЩЕНИЙ ---

@dp.message_handler(content_types=types.ContentType.ANY)
async def handle_all_messages(message: types.Message):
    """Обрабатывает все входящие сообщения, включая медиа."""
    user_id = message.from_user.id
    username = message.from_user.username

    # Обновляем никнейм, если изменился или отсутствует
    current_nickname = user_nicknames.get(user_id)
    new_nickname = username or "" # Используем пустую строку, если ника нет
    if current_nickname != new_nickname:
         user_nicknames[user_id] = new_nickname
         save_data()
         logging.info(f"Обновлен никнейм для {user_id}: {'@' + new_nickname if new_nickname else 'ID'}")

    # Регистрируем пользователя, если его нет в статусах (на случай, если /start не был нажат)
    if user_id not in users_status and user_id not in ADMIN_IDS:
        users_status[user_id] = {"media_status": "?", "review_status": "?", "admin_status": ""}
        logging.info(f"Пользователь {user_id} ({'@' + new_nickname if new_nickname else 'ID'}) автоматически зарегистрирован при получении сообщения.")
        save_data()

    # Проверка на бан (после регистрации/обновления ника)
    if user_id not in ADMIN_IDS and is_user_banned(user_id):
        # Игнорируем сообщения от забаненных пользователей (или отвечаем о бане)
        # await message.reply("Вы заблокированы.")
        logging.info(f"Получено сообщение от забаненного пользователя {user_id}. Игнорируется.")
        return

    # Определяем, является ли сообщение медиа для сохранения
    is_media_to_save = message.photo or message.video or message.document # Добавьте другие типы если нужно

    if user_id not in ADMIN_IDS and is_media_to_save:
        logging.debug(f"Получено медиа сообщение {message.message_id} ({message.content_type}) от {user_id}")

        # --- Логика накопления медиа ---
        if user_id not in user_media_accumulation:
            user_media_accumulation[user_id] = {"files": [], "timer_task": None}

        # Отменяем предыдущий таймер, если он был активен
        if user_media_accumulation[user_id]["timer_task"]:
            user_media_accumulation[user_id]["timer_task"].cancel()
            logging.debug(f"Предыдущий таймер обработки медиа для {user_id} отменен.")

        # Добавляем текущее сообщение в список для обработки
        user_media_accumulation[user_id]["files"].append(message)
        logging.debug(f"Медиа сообщение {message.message_id} добавлено в очередь для {user_id}. Всего в очереди: {len(user_media_accumulation[user_id]['files'])}")

        # Запускаем новый таймер отложенной обработки
        new_task = asyncio.create_task(
             delayed_media_processing(user_id, MEDIA_ACCUMULATION_DELAY)
        )
        user_media_accumulation[user_id]["timer_task"] = new_task
        logging.debug(f"Новый таймер обработки медиа для {user_id} запущен на {MEDIA_ACCUMULATION_DELAY} сек.")
        # --- Конец логики накопления ---

    elif message.text and message.text.startswith('/'):
        # Это команда, она будет обработана другими хэндлерами
        logging.debug(f"Получена команда {message.text} от {user_id}. Ожидается обработка командным хэндлером.")
        pass # Ничего не делаем, позволяем обработчикам команд сработать
    elif message.from_user.id in ADMIN_IDS:
         # Сообщение от админа, не команда и не медиа - возможно, ответ на что-то или просто текст
         # Также проверяем состояние админа (например, ожидание шаблона)
         if user_id in admin_states and admin_states[user_id] == "waiting_for_template":
              # Этот случай обрабатывается хэндлером update_broadcast_template
              pass
         else:
              logging.info(f"Получено текстовое сообщение от админа {user_id}: {message.text[:50]}...")
              # Можно добавить реакцию, если нужно
    else:
        # Сообщение от обычного пользователя, не команда и не медиа
        logging.info(f"Получено не медиа и не команда от пользователя {user_id}: {message.text[:50]}...")
        # Можно ответить пользователю, что бот ожидает команды или медиа
        # await message.reply("Используйте кнопки для взаимодействия или отправьте медиафайлы.")


# --- ЗАПУСК БОТА ---

async def on_startup(dp: Dispatcher):
    """Выполняется при запуске бота."""
    logging.info("Загрузка данных...")
    load_data()
    logging.info("Запуск фоновой задачи обработки очереди сообщений...")
    asyncio.create_task(process_message_queue())
    logging.info("Бот успешно запущен!")

async def on_shutdown(dp: Dispatcher):
    """Выполняется при остановке бота."""
    logging.info("Остановка бота...")
    # Корректное завершение задач (если они есть и их нужно завершать)
    # Сохранение данных перед выходом
    logging.info("Сохранение данных перед выходом...")
    save_data()
    logging.info("Бот остановлен.")


if __name__ == '__main__':
    logging.info("Инициализация и запуск бота...")
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup, on_shutdown=on_shutdown)
